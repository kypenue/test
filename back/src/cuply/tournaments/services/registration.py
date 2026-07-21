""" Services for tournament registration management. """
from datetime import datetime
from io import BytesIO
from typing import Optional
from tempfile import NamedTemporaryFile

import pytz
import openpyxl
from sqlalchemy import case, func
from starlette import status
from fastapi import HTTPException

import constants
from backlib.pagination import AsyncPaginator
from cuply.auth.models import UserModel
from backlib.repo_helpers import raise_not_found_if_none
from cuply.cuply_unit_of_work import AsyncUnitOfWork
from cuply.tournaments.models import TournamentModel, RegistrationStatus, LifecycleTournamentStatus, \
    TournamentPaymentType
from cuply.tournaments.schemas.registration import TournamentRegisteredUserReadSchema
from cuply.tournaments.exceptions.registration import (
    TelegramNotVerifiedRegistrationError,
    UserBlockedRegistrationError,
    RegistrationNotStartedError,
    RegistrationEndedError,
    TooManyParticipantsError,
    AgeLimitRegistrationError,
    UserAlreadyRegisteredError, PlatformNotFoundError,
)
from cuply.tournaments.schemas.registration import (
    TournamentRegisteredUserWriteSchema,
    TournamentSetRegistrationStatusSchema,
)
from cuply.tournaments.filters.registration import TournamentRegisteredUserFilter
from cuply.tournaments.models import TournamentRegisteredUserModel
from cuply.accounts.models import AccountModel
from cuply.tournaments.services.permissions import TournamentPermissionService
from dateutil.relativedelta import relativedelta


class TournamentRegistrationService:
    """ Service for tournament registration management. """
    FREE_TOURNAMENT_MAX_PARTICIPANT_NUMBER = 16

    def __init__(self):
        self.permission_service = TournamentPermissionService()

    async def get_all_tournaments_paginated(
        self,
        uow: AsyncUnitOfWork,
        user_id: int,
        page: int,
        per_page: int,
        order_by: list[str] | None,
        search: str,
        filter_instance: TournamentRegisteredUserFilter,
        user: UserModel,
    ) -> dict:
        """ Get all tournaments that user registered for. """
        paginator = AsyncPaginator(
            session=uow.session,
            model_class=TournamentRegisteredUserModel,
            schema_class=TournamentRegisteredUserReadSchema,
            page=page,
            per_page=per_page,
            order_by=order_by,
            search=search,
            search_by=["name"],
            query=uow.registration_repo.get_filter_by_query(user_id=user_id),
            filter_instance=filter_instance,
        )
        return await paginator.get_result()

    async def get_all_users_paginated(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        page: int,
        per_page: int,
        order_by: list[str] | None,
        search: str,
        filter_instance: TournamentRegisteredUserFilter,
        user: UserModel,
        assigned_to_tournament: Optional[bool],
    ) -> dict:
        """ Get all users registered for tournament. """
        if assigned_to_tournament in [False, True]:
            users = []
            initial_series = await uow.series.find_all(tournament_id=tournament_id, is_initial=True)
            for series in initial_series:
                if series.gamer1:
                    users.append(series.gamer1.user_id)
                if series.gamer2:
                    users.append(series.gamer2.user_id)

            query = uow.registration_repo.get_for_tournament_query(tournament_id, assigned_to_tournament, users)
        else:
            query = uow.registration_repo.get_for_tournament_query(tournament_id)

        sn_exp = func.concat(UserModel.surname, ' ', UserModel.name)
        ns_exp = func.concat(UserModel.name, ' ', UserModel.surname)

        paginator = AsyncPaginator(
            session=uow.session,
            model_class=TournamentRegisteredUserModel,
            schema_class=TournamentRegisteredUserReadSchema,
            page=page,
            per_page=per_page,
            order_by=order_by,
            search=search,
            search_by=[sn_exp, ns_exp, AccountModel.login],
            query=query,
            filter_instance=filter_instance,
            check_fields=False,
        )
        return await paginator.get_result()

    async def get_participants_xlsx(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        user: UserModel,
    ) -> BytesIO:
        """ Get all users registered for tournament report in Excel file. """
        query = uow.registration_repo.get_filter_query(
            TournamentRegisteredUserModel.tournament_id == tournament_id,
        ).order_by(case(
            (TournamentRegisteredUserModel.status == RegistrationStatus.PENDING, 1),
            (TournamentRegisteredUserModel.status == RegistrationStatus.NOT_REGISTERED, 2),
            (TournamentRegisteredUserModel.status == RegistrationStatus.APPROVED, 3),
            (TournamentRegisteredUserModel.status == RegistrationStatus.DENIED, 4),
            else_=5,
        ))
        res = await uow.session.execute(query)
        data = []

        headers = ['ФИО', 'Логин', 'Telegram', 'Дата рождения', 'Статус', 'Дата заявки']
        for registered_user in res.scalars():
            if registered_user.status == RegistrationStatus.PENDING:
                reg_status = 'Ожидает подтверждения'
            elif registered_user.status == RegistrationStatus.NOT_REGISTERED:
                reg_status = 'Не зарегистрирован'
            elif registered_user.status == RegistrationStatus.APPROVED:
                reg_status = 'Зарегистрирован'
            elif registered_user.status == RegistrationStatus.DENIED:
                reg_status = 'Отклонен'
            else:
                reg_status = ''

            data.append(
                {
                    'ФИО': f'{registered_user.account.user.name} {registered_user.account.user.surname}'.strip(),
                    'Логин': registered_user.account.login,
                    'Telegram': registered_user.account.user.tg_login,
                    'Дата рождения': registered_user.account.user.birth_date,
                    'Статус': reg_status,
                    'Дата заявки': registered_user.created_at,
                }
            )

        wb = openpyxl.Workbook()
        ws = wb.active

        for col_num, value in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=value)

        for row_num, row in enumerate(data, start=1):
            for col_num, value in enumerate(row):
                ws.cell(row=row_num + 1, column=col_num, value=value)

        filters = ws.auto_filter

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            output = BytesIO(tmp.read())
            output.seek(0)

        return output

    async def get_registration(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        registration_id: int,
        user: UserModel,
    ) -> TournamentRegisteredUserReadSchema:
        """ Get registration by id. """
        registration = await uow.registration_repo.get_full_participant(
            TournamentRegisteredUserModel.id == registration_id,
            TournamentRegisteredUserModel.tournament_id == tournament_id,
        )
        raise_not_found_if_none(registration, registration_id)
        return TournamentRegisteredUserReadSchema.model_validate(registration)

    async def set_registration_status(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        registration_id: int,
        schema: TournamentSetRegistrationStatusSchema,
        user: UserModel,
    ):
        """ Set registration status for tournament registration. """
        data = schema.model_dump()

        tournament: TournamentModel = await uow.tournament_repo.find_one(id=tournament_id)
        raise_not_found_if_none(tournament, tournament_id)

        await self.permission_service.check_can_set_registration_status(uow, user, tournament=tournament)

        registration_status = data['status']
        if registration_status == RegistrationStatus.APPROVED:
            if tournament.payment_type == TournamentPaymentType.FREE:
                approved_participants_number = await uow.registration_repo.get_approved_registration_count(
                    tournament_id=tournament.id
                )
                if approved_participants_number >= self.FREE_TOURNAMENT_MAX_PARTICIPANT_NUMBER:
                    raise TooManyParticipantsError(
                        f"Бесплатные турниры поддерживают не более {self.FREE_TOURNAMENT_MAX_PARTICIPANT_NUMBER} "
                        f"участников"
                    )

        registration = await uow.registration_repo.get_full_participant(
            TournamentRegisteredUserModel.id==registration_id,
            TournamentRegisteredUserModel.tournament_id==tournament_id,
        )
        raise_not_found_if_none(registration, registration_id)

        await uow.registration_repo.edit_one(registration_id, data)
        await uow.commit()

    async def _check_tournament_exists(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
    ) -> TournamentModel:
        """ Get tournament by id or raise exceptions. """
        tournament = await uow.tournament_repo.find_one(id=tournament_id)
        raise_not_found_if_none(tournament, tournament_id)
        return tournament

    async def _check_account_exists(
        self,
        uow: AsyncUnitOfWork,
        account_id: int,
        user: UserModel,
    ) -> AccountModel:
        """ Get account by id or raise exceptions. """
        account = await uow.account_repo.find_one(id=account_id)
        raise_not_found_if_none(account, account_id)
        if account.user_id != user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=constants.NOT_ENOUGH_PERMISSIONS_MSG,
            )
        return account

    async def _check_registration_limits(
        self,
        uow: AsyncUnitOfWork,
        user: UserModel,
        tournament: TournamentModel,
        account: AccountModel,
    ):
        """ Checks if the user can register for tournament or raise exceptions. """
        now = datetime.utcnow().replace(tzinfo=pytz.utc)

        if not user.tg_login:
            raise TelegramNotVerifiedRegistrationError()

        user_block = await uow.user_block_repo.find_one(user_id=user.id, creator_id=tournament.creator_id)
        if user_block is not None and (user_block.blocked_until is None or user_block.blocked_until >= now):
            raise UserBlockedRegistrationError()

        if now < tournament.registration_start:
            raise RegistrationNotStartedError()

        if now > tournament.registration_end:
            raise RegistrationEndedError()

        approved_registration_count = await uow.registration_repo.get_approved_registration_count(tournament.id)
        if tournament.should_limit_participants and approved_registration_count >= tournament.participants_number:
            raise TooManyParticipantsError()

        player_age = relativedelta(tournament.tournament_end.date(), user.birth_date).years
        if player_age < tournament.min_age:
            raise AgeLimitRegistrationError()

        allowed_platforms = await uow.tournament_platforms.find_all(tournament_id=tournament.id)
        is_platform_found = False
        for allowed_platform in allowed_platforms:
            if allowed_platform.platform_id == account.platform_id:
                is_platform_found = True
                break
        if not is_platform_found:
            raise PlatformNotFoundError()

    async def register(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        schema: TournamentRegisteredUserWriteSchema,
        user: UserModel,
    ) -> TournamentRegisteredUserReadSchema:
        """ Register user for tournament. """
        data = schema.model_dump()
        data["tournament_id"] = tournament_id

        account = await self._check_account_exists(uow, schema.model_dump()["account_id"], user)

        registration = await uow.registration_repo.get_full_participant(
            AccountModel.user_id == user.id, TournamentRegisteredUserModel.tournament_id == tournament_id,
        )

        if registration is not None and data["account_id"] == registration.account_id:
            return TournamentRegisteredUserReadSchema.model_validate(registration)
        elif registration is not None:
            raise UserAlreadyRegisteredError()

        tournament = await self._check_tournament_exists(uow, tournament_id)
        await self._check_registration_limits(uow, user, tournament, account)

        if tournament.lifecycle_status != LifecycleTournamentStatus.REGISTRATION_OPENED:
            if tournament.lifecycle_status == LifecycleTournamentStatus.REGISTRATION_NOT_STARTED:
                raise RegistrationNotStartedError()
            raise RegistrationEndedError()

        data["status"] = RegistrationStatus.APPROVED if tournament.is_auto_approval else RegistrationStatus.PENDING
        registration_id = await uow.registration_repo.add_one(data)
        registration = await uow.registration_repo.get_full_participant(TournamentRegisteredUserModel.id == registration_id)

        await uow.commit()

        return TournamentRegisteredUserReadSchema.model_validate(registration)

    async def unregister(
        self,
        uow: AsyncUnitOfWork,
        tournament_id: int,
        user: UserModel,
    ):
        """ Unregister user from tournament. """
        await self._check_tournament_exists(uow, tournament_id)

        registration = await uow.registration_repo.get_full_participant(
            AccountModel.user_id == user.id, TournamentModel.id == tournament_id,
        )

        tournament = await self._check_tournament_exists(uow, tournament_id)
        if tournament.lifecycle_status != LifecycleTournamentStatus.REGISTRATION_OPENED:
            if tournament.lifecycle_status == LifecycleTournamentStatus.REGISTRATION_NOT_STARTED:
                raise RegistrationNotStartedError()
            raise RegistrationEndedError()

        if registration is not None:
            await uow.registration_repo.delete_one(registration.id)
            await uow.commit()
